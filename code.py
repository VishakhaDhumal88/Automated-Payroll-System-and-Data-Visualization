from tkinter import *
from tkinter import ttk
import tkinter as tk
from openpyxl import load_workbook
from datetime import date
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import webbrowser
from PIL import Image, ImageTk
import os
newline = os.linesep

selected_id = None
selected_name = None
selected_ctc = None
today = None
selected_gender=None
selected_Department=None
selected_GSalary=None
selected_HRA=None
selected_Basic=None
selected_Name=None
selected_Designation=None

def Features():
    print("Function 1 selected")

def Employee_Data():
    print("Function 1 selected")

def Overtime_Analysis():
    print("Function 1 selected")

def Employee_Accounts():
    print("Function 1 selected")

def Help():
    print("Function 1 selected")

def Contact():
    print("Function 1 selected")

def set_hand_cursor(event):
    event.widget.config(cursor="hand2")

def reset_cursor(event):
    event.widget.config(cursor="")

def generate_pdf(slip):
    pdf_name = "Salary Slip.pdf"
    c = canvas.Canvas(pdf_name, pagesize=letter)
    c.setFont("Helvetica", 12)

    lines = slip.split("\n")
    y = 750

    for line in lines:
        c.drawString(100, y, line)
        y -= 14

    # Add image to the PDF
    image_path = "background_image.jpg"  # Replace with the actual path to your image file
    c.drawImage(image_path, 0, 0, width=letter[0], height=letter[1])

    c.save()

    # Open the PDF file
    webbrowser.open(pdf_name)

def verify_name():
    global selected_id, selected_name

    selected_id = var.get()

    if selected_id in id_name_dict:
        selected_name = id_name_dict[selected_id]
        print(f"Employee Name: {selected_name}")
    else:
        print("Invalid Employee ID")

def display_salary_slip(*args):
    global selected_id, selected_name, selected_ctc, today,selected_gender,selected_Department,selected_Designation,selected_Basic,selected_Transport,selected_HRA,selected_GSalary,selected_Name

    selected_id = var.get()

    if selected_id in id_name_dict:
        selected_name = id_name_dict[selected_id]
        selected_ctc = id_ctc_dict[selected_id]
        selected_gender= id_Department_dict[selected_id]
        selected_Department=id_WeeklyOff_dict[selected_id]
        selected_Designation=id_paidHolidays_dict[selected_id]
        selected_Basic=id_Basic_dict[selected_id]
        selected_HRA=id_HRA_dict[selected_id]
        selected_Transport=id_Transport_dict[selected_id]
        selected_GSalary=id_GSalary_dict[selected_id]
        selected_Name=id_Name_dict[selected_id]
        today = date.today().strftime("%B %d, %Y")
    else:
        print("Invalid Employee ID")

def generate_salary_slip():
    global selected_id, selected_name, selected_ctc, today,selected_Department,selected_gender,selected_Designation,selected_HRA,selected_Transport,selected_Basic

    if selected_id is not None and selected_name is not None:
        slip = (
            "                                            CAREERBOOK \n"
            "     \n"
            "----------------------------------------------------------------------------------------------------------------\n"
             "     \n"
            f"Employee ID: {selected_id:>60}\n"
            "     \n"
            f"Empoyee Name:-{selected_Name:>60}\n"
            "        \n"
            f"Father's Name:{selected_gender:>60}\n"
            "     \n"
            f"Salary Date: {today:>60}\n"
            "     \n"
            f"Department:{selected_Department:>60}\n"
            "     \n"
            f"Basic Pay:{selected_Basic:>61}\n"
            "     \n"
            f"HRA:{selected_HRA:>70}\n"
            "     \n"
            f"Transport Alliances:{selected_Transport:>48}\n"
            "     \n"
            f"Gross Salary:{selected_GSalary:>58}\n"
            "     \n"
            f"CTC: Rs. {selected_ctc:>64}\n"
            "     \n"
        )

        generate_pdf(slip)
    else:
        print("Invalid Employee ID or Name")

root = Tk()
root.title('Payroll Slip Generation')
root.geometry("1280x720")

background_image = Image.open("pbl6.jpg")
background_photo = ImageTk.PhotoImage(background_image)
background_label = Label(root, image=background_photo)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

root.configure(bg='white')
root.resizable(False, False)

f1 = tk.Frame(root, width=198, height=563, borderwidth=5, bg="Sky Blue")
f1.pack(side=tk.LEFT, anchor=tk.SW)
f1.pack_propagate(0)

l = tk.Label(f1, text="Dashboard", font=("Arial", 20, "bold"))
l.pack(anchor=tk.NW)

button1 = Button(
    f1, text="Features", command=Features, bg='white', highlightbackground=f1["bg"]
)
button1.pack(side="top", padx=0, pady=10, anchor="w")

button1.bind("<Enter>", set_hand_cursor)
button1.bind("<Leave>", reset_cursor)

button2 = Button(
    f1, text="View Employee Data", command=Employee_Data, bg='grey', highlightbackground=f1["bg"]
)
button2.pack(side="top", padx=0, pady=10, anchor="w")

button2.bind("<Enter>", set_hand_cursor)
button2.bind("<Leave>", reset_cursor)

button3 = Button(
    f1, text="Overtime_Analysis", command=Overtime_Analysis, bg='grey', highlightbackground=f1["bg"]
)
button3.pack(side="top", padx=0, pady=10, anchor="w")

button3.bind("<Enter>", set_hand_cursor)
button3.bind("<Leave>", reset_cursor)

button4 = Button(f1, text="Help", command=Help, bg='grey', highlightbackground=f1["bg"])
button4.pack(side="top", padx=0, pady=10, anchor="w")

button4.bind("<Enter>", set_hand_cursor)
button4.bind("<Leave>", reset_cursor)

button5 = Button(f1, text="Contact", command=Contact, bg='grey', highlightbackground=f1["bg"])
button5.pack(side="top", padx=0, pady=10, anchor="w")

button5.bind("<Enter>", set_hand_cursor)
button5.bind("<Leave>", reset_cursor)

wb = load_workbook('salary.xlsx')
ws = wb.active

column_a = ws['B']
column_b = ws['D']
column_r = ws['AN']
column_s = ws['E']
column_c = ws['L']
column_d = ws['K']
column_w = ws['W']
column_x = ws['AO']
column_y = ws['AP']
column_z = ws['Z']
column_d = ws['D']

id_name_dict = {cell.value: name_cell.value for cell, name_cell in zip(column_a, column_b)}
id_ctc_dict = {cell.value: ctc_cell.value for cell, ctc_cell in zip(column_a, column_r)}
id_Department_dict = {cell.value: Department_cell.value for cell, Department_cell in zip(column_a, column_s)}
id_WeeklyOff_dict = {cell.value: WeeklyOff_cell.value for cell, WeeklyOff_cell in zip(column_a, column_c)}
id_paidHolidays_dict = {cell.value: paidHolidays_cell.value for cell, paidHolidays_cell in zip(column_a, column_d)}
id_GSalary_dict = {cell.value: GSalary_cell.value for cell, GSalary_cell in zip(column_a, column_w)}
id_Basic_dict = {cell.value: Basic_cell.value for cell, Basic_cell in zip(column_a, column_x)}
id_HRA_dict = {cell.value: HRA_cell.value for cell, HRA_cell in zip(column_a, column_y)}
id_Transport_dict = {cell.value: Transport_cell.value for cell, Transport_cell in zip(column_a, column_z)}
id_Name_dict = {cell.value: Name_cell.value for cell, Name_cell in zip(column_a, column_d)}

var = StringVar()
options = list(id_name_dict.keys())

def clear_cb():
    my_combobox.set('')

Label(root, text="AUTOMATED PAYROLL SYSTEM", font=("Times", "20", "bold")).pack(
    padx=40, pady=10
)

my_combobox = ttk.Combobox(root, textvariable=var, values=[*options], state='readonly')
my_combobox.current(1)
my_combobox.pack()

my_combobox.bind("<Enter>", set_hand_cursor)
my_combobox.bind("<Leave>", reset_cursor)

var.trace('w', display_salary_slip)

button = ttk.Button(root, text="Generate PDF", command=generate_salary_slip)
button.pack(pady=10)

button.bind("<Enter>", set_hand_cursor)
button.bind("<Leave>", reset_cursor)

root.mainloop()
